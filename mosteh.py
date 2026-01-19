"""
Скрипт для генерации отчета МосТех

=== КАК ЗАПУСТИТЬ ===

1. С автоматическим поиском CSV в текущей директории:
   python mosteh.py --start_date 2025-01-15 --end_date 2025-01-15

2. С указанием входного и выходного файла:
   python mosteh.py --input data.csv --output report.xlsx --start_date 2025-01-15 --end_date 2025-01-15

Форматы:
  --start_date и --end_date: ГГГГ-ММ-ДД (например: 2025-01-15)
  Скрипт автоматически подберет год из данных, если это необходимо
"""

import pandas as pd
import argparse
from datetime import datetime
import re
import openpyxl
import os
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font

def find_csv_file(directory):
    """Находит первый CSV файл в указанной директории"""
    for file in os.listdir(directory):
        if file.lower().endswith('.csv'):
            return os.path.join(directory, file)
    return None

def parse_arguments():
    parser = argparse.ArgumentParser(description='Скрипт для генерации отчета для МосТех')
    parser.add_argument('--input', type=str, help='Путь к входному CSV файлу (если не указан, ищется в текущей директории)')
    parser.add_argument('--output', type=str, help='Путь к выходному XLSX файлу (по умолчанию report_<дата>.xlsx в текущей директории)')
    parser.add_argument('--start_date', type=str, required=True, help='Начальная дата в формате ГГГГ-ММ-ДД')
    parser.add_argument('--end_date', type=str, required=True, help='Конечная дата в формате ГГГГ-ММ-ДД')
    return parser.parse_args()

def format_date_for_excel(date_str):
    """Преобразует дату в формат ММ/ДД/ГГ ЧЧ:ММ для Excel"""
    try:
        # Пытаемся распарсить дату из формата '2026-01-15 20:21:51'
        dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        # Форматируем в нужный формат ММ/ДД/ГГ ЧЧ:ММ
        return dt.strftime('%m/%d/%y %H:%M')
    except:
        return date_str

def clean_parameters(param_str):
    """Очищает и форматирует поле parameters для отображения в Excel"""
    if pd.isna(param_str) or param_str == "[]":
        return ""
    
    # Заменяем возможные варианты разделителей
    param_str = param_str.replace(";", ",").replace("{", "[").replace("}", "]")
    
    # Заменяем кавычки и приводим к единому формату
    param_str = param_str.replace("'", '"').replace('="', "='").replace('="', "='")
    
    # Упрощаем форматирование для лучшего отображения в Excel
    if param_str.startswith("[{") and param_str.endswith("}]"):
        param_str = param_str[2:-2]
        parts = param_str.split("}, {")
        cleaned_parts = []
        for part in parts:
            cleaned_part = re.sub(r'\s+', ' ', part.strip())
            if cleaned_part:
                cleaned_parts.append(cleaned_part)
        return "], [".join(["{" + p + "}" for p in cleaned_parts])
    
    return param_str

def try_read_csv(file_path, encodings=None):
    """Пытается прочитать CSV файл с разными кодировками"""
    if encodings is None:
        encodings = ['utf-8-sig', 'cp1251', 'cp866', 'iso-8859-5', 'utf-16', 'windows-1252']
    
    for encoding in encodings:
        try:
            print(f"Попытка чтения с кодировкой {encoding}...")
            df = pd.read_csv(file_path, sep=';', encoding=encoding, on_bad_lines='warn')
            print(f"Успешно загружено с кодировкой {encoding}")
            return df, encoding
        except Exception as e:
            print(f"Не удалось загрузить с кодировкой {encoding}: {str(e)}")
    
    raise ValueError("Не удалось прочитать файл ни с одной из доступных кодировок")

def generate_report(input_file, output_file, start_date, end_date):
    print(f"Начало обработки данных. Диапазон дат: {start_date} - {end_date}")
    
    # Пытаемся прочитать файл с разными кодировками
    try:
        df, used_encoding = try_read_csv(input_file)
        print(f"Файл успешно загружен с кодировкой: {used_encoding}")
        print(f"Загружено {len(df)} записей из файла")
    except Exception as e:
        print(f"Критическая ошибка при чтении файла {input_file}:")
        print(str(e))
        print("\nДоступные кодировки: 'utf-8-sig', 'cp1251', 'cp866', 'iso-8859-5', 'utf-16', 'windows-1252'")
        print("Попробуйте изменить кодировку файла вручную и повторить попытку.")
        return
            
    
    # Фильтрация по "МосТех" в названии партнера
    df = df[df['partner.name'].str.contains('МосТех', case=False, na=False)]
    print(f"После фильтрации по 'МосТех': {len(df)} записей")
    
    # Преобразуем даты для фильтрации
    print(f"\nПервые 5 дат в payment_time:")
    print(df['payment_time'].head())
    
    # Пробуем разные форматы дат
    date_formats = [
        '%Y-%m-%d %H:%M:%S',  # 2025-01-15 10:30:00
        '%d.%m.%Y %H:%M',     # 15.01.2025 10:30
        '%Y-%m-%dT%H:%M:%S',  # 2025-01-15T10:30:00
        '%d.%m.%Y',           # 15.01.2025
        '%Y-%m-%d'            # 2025-01-15
    ]
    
    # Пытаемся преобразовать даты разными способами
    for fmt in date_formats:
        try:
            df['payment_time_dt'] = pd.to_datetime(df['payment_time'], format=fmt, errors='coerce')
            if not df['payment_time_dt'].isna().all():
                print(f"\nУспешное преобразование дат с форматом: {fmt}")
                print("Примеры преобразованных дат:")
                print(df[['payment_time', 'payment_time_dt']].head())
                break
        except Exception as e:
            print(f"Ошибка при преобразовании с форматом {fmt}: {str(e)}")
    else:
        print("\nНе удалось определить формат даты. Используем автоматическое определение.")
        df['payment_time_dt'] = pd.to_datetime(df['payment_time'], errors='coerce')
    
    # Определяем год из данных
    if not df.empty and 'payment_time_dt' in df.columns and not df['payment_time_dt'].isna().all():
        data_year = df['payment_time_dt'].dt.year.dropna().iloc[0]
        print(f"\nГод в данных: {data_year}")
        
        # Обновляем год в датах фильтрации, если он не совпадает
        start_date_parts = start_date.split('-')
        end_date_parts = end_date.split('-')
        
        if len(start_date_parts) == 3 and int(start_date_parts[0]) != data_year:
            start_date = f"{data_year}-{start_date_parts[1]}-{start_date_parts[2]}"
            end_date = f"{data_year}-{end_date_parts[1]}-{end_date_parts[2]}"
            print(f"Обновлены даты фильтрации: с {start_date} по {end_date}")
    
    # Преобразуем граничные даты в нужный формат
    start_datetime = pd.to_datetime(start_date)
    end_datetime = pd.to_datetime(end_date) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    
    print(f"\nФильтрация по датам с {start_datetime} по {end_datetime}")
    
    # Фильтруем по дате оплаты (payment_time)
    df_filtered = df[
        (df['payment_time_dt'].notna()) & 
        (df['payment_time_dt'] >= start_datetime) & 
        (df['payment_time_dt'] <= end_datetime)
    ]
    
    # Дополнительная отладочная информация
    if not df_filtered.empty:
        print("\nПримеры отфильтрованных записей:")
        print(df_filtered[['payment_time', 'payment_time_dt']].head())
    
    print(f"После фильтрации по датам: {len(df_filtered)} записей из {len(df)}")
    
    # Если после фильтрации нет записей, выводим диапазон дат в данных
    if len(df_filtered) == 0 and len(df) > 0:
        print("\nДиапазон дат в данных:")
        print(f"Минимальная дата: {df['payment_time_dt'].min()}")
        print(f"Максимальная дата: {df['payment_time_dt'].max()}")
    
    df = df_filtered
    
    # Выбираем только нужные колонки в правильном порядке
    required_columns = [
        'id', 'partner.id', 'partner.name', 'pid', 'status',
        'phone', 'amount', 'created', 'changed',
        'payment_time', 'parameters'
    ]
    
    # Создаем новый DataFrame только с нужными колонками
    report_df = df[required_columns].copy()
    
    # Форматируем даты для Excel
    report_df['created'] = report_df['created'].apply(format_date_for_excel)
    report_df['changed'] = report_df['changed'].apply(format_date_for_excel)
    report_df['payment_time'] = report_df['payment_time'].apply(format_date_for_excel)
    
    # Очищаем поле parameters
    report_df['parameters'] = report_df['parameters'].apply(clean_parameters)
    
    # Сохраняем в Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name='Лист1')
    
    # Применяем стилизацию к Excel файлу
    apply_excel_formatting(output_file)
    
    print(f"Отчет успешно сгенерирован и сохранен в {output_file}")
    print(f"В отчет включено {len(report_df)} записей")

def apply_excel_formatting(output_file):
    """Применяет форматирование к Excel файлу для соответствия образцу"""
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Лист1']
    
    # Задаем стили границ
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Применяем заголовки в первой строке
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    
    # Применяем границы ко всем ячейкам
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
    
    # Настраиваем ширину колонок для лучшей читаемости
    column_widths = {
        'A': 10,  # id
        'B': 10,  # partner.id
        'C': 40,  # partner.name
        'D': 15,  # pid
        'E': 12,  # status
        'F': 15,  # phone
        'G': 12,  # amount
        'H': 18,  # created
        'I': 18,  # changed
        'J': 18,  # payment_time
        'K': 70   # parameters
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    wb.save(output_file)

def main():
    args = parse_arguments()
    
    # Проверяем корректность формата дат
    try:
        datetime.strptime(args.start_date, '%Y-%m-%d')
        datetime.strptime(args.end_date, '%Y-%m-%d')
    except ValueError:
        print("Ошибка: даты должны быть в формате ГГГГ-ММ-ДД")
        return
    
    # Определяем входной файл
    input_file = args.input
    if not input_file:
        print("Входной файл не указан, ищем CSV в текущей директории...")
        input_file = find_csv_file(os.getcwd())
        if not input_file:
            print("Ошибка: не найден CSV файл в текущей директории")
            return
        print(f"Найден файл: {input_file}")
    
    # Определяем выходной файл
    output_file = args.output
    if not output_file:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(os.getcwd(), f'report_{timestamp}.xlsx')
        print(f"Выходной файл не указан, будет создан: {output_file}")
    
    generate_report(
        input_file=input_file,
        output_file=output_file,
        start_date=args.start_date,
        end_date=args.end_date
    )

if __name__ == "__main__":
    main()